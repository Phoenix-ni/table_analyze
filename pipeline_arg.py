import cv2
import numpy as np
import json
from sklearn.cluster import DBSCAN
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from ultralytics import YOLO
from paddleocr import PaddleOCR
from openai import OpenAI
import pandas as pd
import os
import re
from datetime import datetime
import argparse

# === 1. 提取表格线条 ===
def extract_table_lines(img_path):
    img = cv2.imread(img_path, 0)
    _, binary = cv2.threshold(img, 128, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
    h_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, h_kernel)
    v_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, v_kernel)
    table_lines = cv2.add(h_lines, v_lines)
    table_lines = cv2.dilate(table_lines, np.ones((2, 2), np.uint8), iterations=1)
    clean_table = 255 * np.ones_like(img)
    clean_table[table_lines > 0] = 0
    return clean_table


# === 2. 生成空 Excel 框架 ===
def create_empty_excel_skeleton(img_path, output_excel_path="empty_table_skeleton.xlsx"):
    clean_table = extract_table_lines(img_path)
    _, binary = cv2.threshold(clean_table, 128, 255, cv2.THRESH_BINARY_INV)
    lines = cv2.HoughLinesP(binary, 1, np.pi / 180, threshold=100, minLineLength=50, maxLineGap=10)
    if lines is None:
        raise ValueError("❌ 未检测到线条")

    h_lines, v_lines = [], []
    for l in lines:
        x1, y1, x2, y2 = l[0]
        if abs(y1 - y2) < 10:
            h_lines.append((x1, y1, x2, y2))
        elif abs(x1 - x2) < 10:
            v_lines.append((x1, y1, x2, y2))

    print(f"检测到水平线 {len(h_lines)} 条，垂直线 {len(v_lines)} 条")

    points = [(v[0], h[1]) for h in h_lines for v in v_lines]
    points = np.array(points)
    clustering = DBSCAN(eps=10, min_samples=1).fit(points)
    labels = clustering.labels_

    clustered_points = []
    for lbl in set(labels):
        cluster = points[labels == lbl]
        clustered_points.append(np.mean(cluster, axis=0))
    clustered_points = np.array(clustered_points)

    grid_x = sorted(np.unique(np.round(clustered_points[:, 0]).astype(int)))
    grid_y = sorted(np.unique(np.round(clustered_points[:, 1]).astype(int)))
    rows, cols = len(grid_y) - 1, len(grid_x) - 1
    print(f"✅ 检测到 {rows} 行 {cols} 列")

    def has_horizontal_line(y, x1, x2):
        segment = binary[max(y - 1, 0):min(y + 1, binary.shape[0]), x1:x2]
        return np.sum(segment) > (x2 - x1) * 255 * 0.3

    def has_vertical_line(x, y1, y2):
        segment = binary[y1:y2, max(x - 1, 0):min(x + 1, binary.shape[1])]
        return np.sum(segment) > (y2 - y1) * 255 * 0.3

    wb = Workbook()
    ws = wb.active
    ws.title = "ReconstructedTable"

    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15
    for r in range(1, rows + 1):
        ws.row_dimensions[r].height = 20

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    visited = np.zeros((rows, cols), dtype=bool)
    merge_info, merge_ranges = {}, []

    for i in range(rows):
        for j in range(cols):
            if visited[i, j]:
                continue
            x1, y1 = grid_x[j], grid_y[i]
            x2, y2 = grid_x[j + 1], grid_y[i + 1]
            rowspan, colspan = 1, 1

            while j + colspan < cols and not has_vertical_line(grid_x[j + colspan], y1, y2):
                colspan += 1
            while i + rowspan < rows and not has_horizontal_line(grid_y[i + rowspan], x1, x2):
                rowspan += 1

            visited[i:i + rowspan, j:j + colspan] = True
            start_cell = f"{get_column_letter(j + 1)}{i + 1}"
            end_cell = f"{get_column_letter(j + colspan)}{i + rowspan}"

            if rowspan > 1 or colspan > 1:
                merge_range = f"{start_cell}:{end_cell}"
                ws.merge_cells(merge_range)
                merge_ranges.append(merge_range)
                for r in range(i, i + rowspan):
                    for c in range(j, j + colspan):
                        merge_info[(r, c)] = (i, j)

            for r in range(i + 1, i + rowspan + 1):
                for c in range(j + 1, j + colspan + 1):
                    ws.cell(r, c).border = border

    wb.save(output_excel_path)
    print(f"✅ 已生成空的表格框架：{output_excel_path}")
    return grid_x, grid_y, rows, cols, merge_info, merge_ranges


# === 3. YOLO 检测单元格 ===
def run_yolo(model_path, img_path, conf=0.25):
    model = YOLO(model_path)
    results = model.predict(source=img_path, conf=conf, verbose=False)
    yolo_boxes = []
    for r in results:
        for box in r.boxes.xyxy.cpu().numpy():
            x1, y1, x2, y2 = map(int, box[:4])
            yolo_boxes.append((x1, y1, x2, y2))
    return yolo_boxes


# === 4. YOLO → Grid 映射 ===
def map_yolo_to_grid(yolo_boxes, grid_x, grid_y):
    rows, cols = len(grid_y) - 1, len(grid_x) - 1
    grid_to_yolo = {}
    grid_centers = [((grid_x[j] + grid_x[j+1]) // 2, (grid_y[i] + grid_y[i+1]) // 2, i, j)
                    for i in range(rows) for j in range(cols)]
    for (x1, y1, x2, y2) in yolo_boxes:
        cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
        gx, gy, i, j = min(grid_centers, key=lambda g: (g[0]-cx)**2 + (g[1]-cy)**2)
        grid_to_yolo[(i, j)] = (x1, y1, x2, y2)
    return grid_to_yolo


# === 5. OCR ===
def save_ocr_result_to_json(result, json_path):
    def default_converter(o):
        if isinstance(o, np.ndarray):
            return o.tolist()
        if isinstance(o, (np.float32, np.float64)):
            return float(o)
        if isinstance(o, (np.int32, np.int64)):
            return int(o)
        return str(o)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result[0], f, ensure_ascii=False, indent=2, default=default_converter)


def extract_text_from_json(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if "rec_texts" in data:
        return " ".join(data["rec_texts"])
    return ""


def fill_excel_with_ocr_content(empty_excel_path, output_excel_path, grid_to_yolo, merge_info, merge_ranges, img_path, ocr_engine):
    wb = load_workbook(empty_excel_path)
    ws = wb.active
    print("跳过重新合并单元格（Excel中已存在合并信息）")

    original_img = cv2.imread(img_path)
    filled_merge_cells = set()
    debug_img = original_img.copy()

    for (i, j), yolo_box in grid_to_yolo.items():
        if (i, j) in merge_info:
            top_left_i, top_left_j = merge_info[(i, j)]
        else:
            top_left_i, top_left_j = i, j

        if (top_left_i, top_left_j) in filled_merge_cells:
            continue

        filled_merge_cells.add((top_left_i, top_left_j))
        x1, y1, x2, y2 = yolo_box
        crop = original_img[y1:y2, x1:x2]

        result = ocr_engine.predict(crop)
        save_ocr_result_to_json(result, "ocr_result.json")
        text = extract_text_from_json("ocr_result.json")

        cell_address = f"{get_column_letter(top_left_j + 1)}{top_left_i + 1}"
        try:
            ws[cell_address] = text
            print(f"填写单元格 ({top_left_i},{top_left_j}) - {cell_address}: {text}")
        except Exception as e:
            print(f"错误: 无法写入单元格 {cell_address}: {e}")

        cv2.rectangle(debug_img, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv2.putText(debug_img, f"{i},{j}", (x1 + 5, y1 + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 0, 0), 2)

    cv2.imwrite("debug_yolo_to_grid.jpg", debug_img)
    wb.save(output_excel_path)
    print(f"✅ 已生成包含内容的 Excel 表格：{output_excel_path}")


# === 6. LLM接入 ===
def ask_LLM(prompt: str):
    client = OpenAI(
        api_key="*****************************",  # 请在此处填写你的API Key
        base_url="******************************"  # 请在此处填写你的API地址（如果有的话）
    )
    response = client.chat.completions.create(
        model="LongCat-Flash-Chat",
        messages=[
            {"role": "system", "content": "你是一个能分析Excel数据并生成Python代码的助手。"},
            {"role": "user", "content": prompt}
        ],
        temperature=0.2
    )
    return response.choices[0].message.content.strip()


def analyze_excel_with_LLM(filepath: str, question: str):
    df = pd.read_excel(filepath)
    prompt = f"""
    我有一个Excel表格，内容如下（前5行）：
    {df.head().to_string(index=False)}

    用户问题是：{question}

    请输出：
    1. 分析思路
    2. 一段可执行的 pandas 代码（不包含导入语句）
    3. 最终答案（如果可直接得出）

    输出格式：
    步骤：
    <说明>
    代码：
    ```python
    <代码>
    ```
    """

    reply = ask_LLM(prompt)
    print("\n🔍 GPT 输出：\n", reply)

    code_match = re.search(r"```python(.*?)```", reply, re.S)
    if code_match:
        code = code_match.group(1).strip()
        print("\n🚀 正在执行生成的代码...\n")
        local_vars = {"df": df}
        try:
            exec(code, {}, local_vars)
            if "result" in local_vars:
                print("✅ 结果:", local_vars["result"])
        except Exception as e:
            print("❌ 执行出错:", e)
    else:
        print("⚠️ 未找到可执行代码。")


# === 7. 主程序 ===
def main():
    parser = argparse.ArgumentParser(description="表格识别 + GPT 分析流水线")
    parser.add_argument("--img", required=True, help="输入图片路径")
    parser.add_argument("--model", required=True, help="YOLO 模型路径")
    parser.add_argument("--question", required=False, help="要问LLM的问题")
    args = parser.parse_args()
    
    img_path = args.img
    model_path = args.model
    question = args.question

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    empty_excel_path = f"empty_table_skeleton_{timestamp}.xlsx"
    final_excel_path = f"final_table_with_content_{timestamp}.xlsx"

    print("步骤1: 生成空的 Excel 表格框架...")
    grid_x, grid_y, rows, cols, merge_info, merge_ranges = create_empty_excel_skeleton(img_path, empty_excel_path)

    print("步骤2: YOLO检测单元格...")
    yolo_boxes = run_yolo(model_path, img_path, conf=0.25)

    print("步骤3: 将YOLO框映射到网格...")
    grid_to_yolo = map_yolo_to_grid(yolo_boxes, grid_x, grid_y)

    print("步骤4: OCR识别并填写内容...")
    ocr_engine = PaddleOCR(use_textline_orientation=True, lang="ch")
    fill_excel_with_ocr_content(empty_excel_path, final_excel_path, grid_to_yolo, merge_info, merge_ranges, img_path, ocr_engine)

    os.remove("ocr_result.json")

    print("🎉 表格识别完成！")
    print(f" - 最终表格: {final_excel_path}")
    # === 新增 LLM 分析 ===
    if question:
        print("步骤5: LLM分析表格内容...")
        analyze_excel_with_LLM(final_excel_path, question)
    else:
        print("💡 提示：如需LLM分析，请提供 --question 参数")


if __name__ == "__main__":
    main()
